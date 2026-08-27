"""PDF and Excel export of a worker's own income/expense + work report."""
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


def build_worker_pdf_report(worker, transactions, work_entries, totals):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=18 * mm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(f"Worker Report — {worker.full_name}", styles["Title"]),
        Paragraph(f"Worker type: {worker.worker_type or '-'} | Generated: {date.today().isoformat()}", styles["Normal"]),
        Spacer(1, 10),
        Paragraph(
            f"Total Income: ₹{totals['income']:,.0f} &nbsp;&nbsp; "
            f"Total Expense: ₹{totals['expense']:,.0f} &nbsp;&nbsp; "
            f"Net: ₹{totals['net']:,.0f}",
            styles["Heading3"],
        ),
        Spacer(1, 12),
        Paragraph("Income & Expense Entries", styles["Heading2"]),
    ]

    tx_rows = [["Date", "Type", "Category", "Amount", "Notes"]]
    for t in transactions:
        tx_rows.append([t.transaction_date.isoformat(), t.kind, t.category, f"{t.amount:,.0f}", t.notes or ""])
    tx_table = Table(tx_rows, repeatRows=1, colWidths=[70, 55, 100, 70, 140])
    tx_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2e7d32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    story.append(tx_table)
    story.append(Spacer(1, 16))

    story.append(Paragraph("Work Entries", styles["Heading2"]))
    work_rows = [["Date", "Description", "Location", "Hours"]]
    for w in work_entries:
        work_rows.append([w.work_date.isoformat(), w.description, w.location or "", str(w.hours_worked or "")])
    work_table = Table(work_rows, repeatRows=1, colWidths=[70, 210, 100, 55])
    work_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1565c0")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
    ]))
    story.append(work_table)

    doc.build(story)
    buf.seek(0)
    return buf


def build_worker_excel_report(worker, transactions, work_entries, totals):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["Worker", worker.full_name])
    ws1.append(["Worker Type", worker.worker_type or ""])
    ws1.append(["Total Income", totals["income"]])
    ws1.append(["Total Expense", totals["expense"]])
    ws1.append(["Net Balance", totals["net"]])
    for row in ws1.iter_rows(min_row=1, max_row=5, min_col=1, max_col=1):
        for cell in row:
            cell.font = Font(bold=True)

    ws2 = wb.create_sheet("Transactions")
    ws2.append(["Date", "Type", "Category", "Amount", "Notes"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for t in transactions:
        ws2.append([t.transaction_date.isoformat(), t.kind, t.category, t.amount, t.notes or ""])

    ws3 = wb.create_sheet("Work Entries")
    ws3.append(["Date", "Description", "Location", "Hours"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    for w in work_entries:
        ws3.append([w.work_date.isoformat(), w.description, w.location or "", w.hours_worked or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
